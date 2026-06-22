import streamlit as st
import math
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

st.set_page_config(page_title="MedOil SC", page_icon="📈", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<link rel="stylesheet"
      href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
<style>
.bi { display:inline-block; vertical-align:-.125em; line-height:1; }
.bi-sm  { font-size:.85rem; }
.bi-md  { font-size:1.1rem; }
.bi-lg  { font-size:1.4rem; }
.bi-xl  { font-size:1.9rem; }
.bi-2x  { font-size:2.2rem; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;500;600;700&family=DM+Mono:wght@400;500&family=DM+Serif+Display:ital@0;1&display=swap');

:root {
    --green-dark:   #1a3d1a;
    --green-mid:    #2d6a2d;
    --green-light:  #4a8c3f;
    --green-pale:   #c8e6c9;
    --gold:         #f5c518;
    --gold-dark:    #c9a000;
    --gold-mid:     #e8b400;
    --bg-page:      #f4f4ee;
    --bg-card:      #ffffff;
    --bg-sidebar:   #152415;
    --text-dark:    #1a2e1a;
    --text-mid:     #3d5c3d;
    --text-muted:   #7a9a7a;
    --border:       #d4e0d4;
}

html, body, [class*="css"] {
    font-family: 'Syne', sans-serif;
    background-color: var(--bg-page) !important;
    color: var(--text-dark) !important;
}

[data-testid="stSidebar"] {
    background-color: var(--bg-sidebar) !important;
    border-right: 1px solid #2d4a2d !important;
}
[data-testid="stSidebar"] * { color: #c8e6c9 !important; }
[data-testid="stSidebar"] h1 { color: #ffffff !important; font-size: 1.2rem !important; }
[data-testid="stSidebar"] hr { border-color: #2d4a2d !important; }

[data-testid="stSidebarNavLink"] {
    border-radius: 8px !important;
    color: #a5c8a5 !important;
    font-size: .88rem !important;
    padding: .5rem .9rem !important;
    margin: 2px 0 !important;
    transition: background .15s;
}
[data-testid="stSidebarNavLink"]:hover {
    background: rgba(245,197,24,.12) !important;
    color: var(--gold) !important;
}
[data-testid="stSidebarNavLink"][aria-current="page"] {
    background: rgba(245,197,24,.18) !important;
    color: var(--gold) !important;
    border-left: 3px solid var(--gold) !important;
    font-weight: 600 !important;
}

.main .block-container {
    background-color: var(--bg-page) !important;
    padding-top: 1.5rem !important;
}

h1, h2, h3, h4 { color: var(--text-dark) !important; }

.formula-box {
    background: #eef4ee;
    border-left: 3px solid var(--green-light);
    border-radius: 8px;
    padding: .9rem 1.1rem;
    font-family: 'DM Mono', monospace;
    font-size: .82rem;
    color: var(--text-mid);
    margin-bottom: 1rem;
    line-height: 1.7;
    border: 1px solid var(--border);
}
.f-eq { color: var(--green-mid); font-weight: 600; }

.alert-critical {
    background: rgba(220,53,53,.07);
    border: 1px solid rgba(220,53,53,.3);
    border-left: 4px solid #dc3535;
    border-radius: 10px;
    padding: 1rem 1.2rem;
    margin-bottom: .7rem;
    color: #8b1a1a;
    display:flex; align-items:flex-start; gap:.7rem;
}
.alert-critical .alert-icon { color:#dc3535; font-size:1.15rem; margin-top:2px; flex-shrink:0; }
.alert-warning {
    background: rgba(245,197,24,.1);
    border: 1px solid rgba(200,160,0,.3);
    border-left: 4px solid var(--gold-dark);
    border-radius: 10px;
    padding: 1rem 1.2rem;
    margin-bottom: .7rem;
    color: #7a5c00;
    display:flex; align-items:flex-start; gap:.7rem;
}
.alert-warning .alert-icon { color:#c9a000; font-size:1.15rem; margin-top:2px; flex-shrink:0; }
.alert-ok {
    background: rgba(74,140,63,.08);
    border: 1px solid rgba(74,140,63,.3);
    border-left: 4px solid var(--green-light);
    border-radius: 10px;
    padding: 1rem 1.2rem;
    margin-bottom: .7rem;
    color: var(--green-mid);
    display:flex; align-items:flex-start; gap:.7rem;
}
.alert-ok .alert-icon { color:var(--green-light); font-size:1.15rem; margin-top:2px; flex-shrink:0; }
.alert-info {
    background: rgba(45,106,45,.07);
    border: 1px solid rgba(45,106,45,.25);
    border-left: 4px solid var(--green-mid);
    border-radius: 10px;
    padding: 1rem 1.2rem;
    margin-bottom: .7rem;
    color: var(--text-mid);
    display:flex; align-items:flex-start; gap:.7rem;
}
.alert-info .alert-icon { color:var(--green-mid); font-size:1.15rem; margin-top:2px; flex-shrink:0; }

.stSelectbox > div > div,
.stNumberInput > div > div > input,
.stTextInput > div > div > input {
    background: #ffffff !important;
    border: 1px solid var(--border) !important;
    color: var(--text-dark) !important;
    border-radius: 8px !important;
}

[data-testid="stWidgetLabel"] p { color: var(--text-mid) !important; font-size: .85rem !important; }

.stButton > button, .stDownloadButton > button {
    background: var(--gold) !important;
    color: var(--green-dark) !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 700 !important;
    font-family: 'Syne', sans-serif !important;
    padding: .55rem 1.4rem !important;
    transition: background .15s, transform .1s !important;
}
.stButton > button:hover, .stDownloadButton > button:hover {
    background: var(--gold-dark) !important;
    transform: translateY(-1px) !important;
}

[data-testid="stTabs"] [role="tablist"] {
    border-bottom: 2px solid var(--border) !important;
    gap: .5rem !important;
}
[data-testid="stTabs"] [role="tab"] {
    color: var(--text-muted) !important;
    background: transparent !important;
    border-radius: 8px 8px 0 0 !important;
    padding: .5rem 1.1rem !important;
    font-weight: 500 !important;
}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] {
    color: var(--green-dark) !important;
    border-bottom: 3px solid var(--gold) !important;
    font-weight: 700 !important;
}

[data-testid="stDataFrame"] {
    border: 1px solid var(--border) !important;
    border-radius: 10px !important;
    overflow: hidden !important;
}

[data-testid="stExpander"] {
    background: #ffffff !important;
    border: 1px solid var(--border) !important;
    border-radius: 10px !important;
}

.stInfo    { background: rgba(45,106,45,.07) !important; border-color: var(--green-light) !important; color: var(--green-mid) !important; }
.stSuccess { background: rgba(74,140,63,.08) !important; border-color: var(--green-light) !important; }
.stWarning { background: rgba(245,197,24,.1) !important; border-color: var(--gold-dark) !important; }
hr { border-color: var(--border) !important; }

[data-testid="stMultiSelect"] span {
    background: var(--green-pale) !important;
    color: var(--green-dark) !important;
    border-radius: 4px !important;
}

/* Source selector pills */
.source-pills {
    display: flex;
    gap: .5rem;
    margin-bottom: 1rem;
}
.source-pill {
    display: inline-flex;
    align-items: center;
    gap: .4rem;
    padding: .45rem 1rem;
    border-radius: 20px;
    font-size: .82rem;
    font-weight: 600;
    border: 2px solid transparent;
    cursor: pointer;
}
.source-pill-active {
    background: var(--green-mid);
    color: #fff;
    border-color: var(--green-mid);
}
.source-pill-inactive {
    background: #fff;
    color: var(--text-mid);
    border-color: var(--border);
}
</style>""", unsafe_allow_html=True)

# ── Plotly theme ──────────────────────────────────────────────────────────────
def light_layout():
    return dict(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(248,250,248,0.6)",
        font=dict(color="#3d5c3d", family="Syne"),
        legend=dict(bgcolor="rgba(255,255,255,.9)", bordercolor="#d4e0d4", borderwidth=1),
        margin=dict(t=20, b=10, l=0, r=0)
    )

C_GREEN  = "#2d6a2d"
C_GOLD   = "#f5c518"
C_GREEN2 = "#4a8c3f"
C_RED    = "#dc3535"
C_PURPLE = "#7a5c9a"
C_MUTED  = "#7a9a7a"

def fmt(n, d=1):
    if n is None or (isinstance(n, float) and (math.isnan(n) or math.isinf(n))): return "—"
    return f"{n:,.{d}f}"

def fmtInt(n):
    if n is None or (isinstance(n, float) and (math.isnan(n) or math.isinf(n))): return "—"
    return f"{round(n):,}"

def mcard(label, value, unit, delta=None, color="#2d6a2d", icon="bi-bar-chart-fill"):
    dh = f"<div style='font-size:.7rem;color:{color};margin-top:4px'>{delta}</div>" if delta else ""
    return f"""<div style='background:#ffffff;border:1px solid #d4e0d4;border-radius:12px;
        padding:1rem 1.2rem;text-align:center;border-top:3px solid {color};
        box-shadow:0 2px 8px rgba(26,61,26,.06)'>
        <div style='font-size:1.3rem;color:{color};margin-bottom:.3rem'>
            <i class="bi {icon}"></i></div>
        <div style='font-size:.62rem;color:#7a9a7a;text-transform:uppercase;letter-spacing:.1em;margin-bottom:.4rem'>{label}</div>
        <div style='font-family:"DM Mono",monospace;font-size:1.6rem;font-weight:500;color:#1a2e1a'>{value}</div>
        <div style='font-size:.68rem;color:#7a9a7a;margin-top:3px'>{unit}</div>{dh}</div>"""

def fbox(title, formula, details, icon="bi-calculator"):
    return (f"<div class='formula-box'>"
            f"<div style='font-size:.62rem;color:#7a9a7a;text-transform:uppercase;letter-spacing:.1em;margin-bottom:6px'>"
            f"<i class='bi {icon}' style='margin-right:5px;color:#4a8c3f'></i>{title}</div>"
            f"<span class='f-eq'>{formula}</span><br/>"
            f"<span style='color:#7a9a7a'>{details}</span></div>")

SOURCE_DELAIS = {"Export": 4 * 30, "Local": 14, "BM": 3}

def get_lt_jours(source):
    return SOURCE_DELAIS.get(source, 30)

def calc_ss(sigma_d, lt_j, z=1.65):
    return round(z * sigma_d * math.sqrt(lt_j))

def calc_pr(d_j, lt_j, ss):
    return round(d_j * lt_j + ss)

def calc_eoq(conso_an, cout_unit, sc_cost=150, taux=0.20):
    if cout_unit <= 0 or conso_an <= 0: return 0, 0, 0
    h   = taux * cout_unit
    eoq = math.sqrt(2 * conso_an * sc_cost / h)
    nb  = conso_an / eoq
    ct  = nb * sc_cost + (eoq / 2) * h
    return round(eoq), round(nb, 1), round(ct, 2)

Z_PAR_CLASSE  = {"A": 2.33, "B": 1.65, "C": 1.28}
NS_PAR_CLASSE = {"A": "99%", "B": "95%", "C": "90%"}

def build_excel_indicateurs(df_res):
    wb   = Workbook()
    thin = Side(style="thin", color="BFBFBF")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(cell, val, bg):
        cell.value = val
        cell.font  = Font(bold=True, size=9, name="Arial")
        cell.fill  = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = brd

    def dat(cell, val, nf=None):
        cell.value = val
        cell.font  = Font(size=9, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd
        if nf: cell.number_format = nf

    ws = wb.active
    ws.title = "Indicateurs SC"

    headers = [
        ("Code article","C8E6C9"),("Description Article","C8E6C9"),
        ("Source","FFF9C4"),("Délai (j)","FFF9C4"),
        ("Classe ABC","DCEDC8"),("Classe XYZ","DCEDC8"),("Classe","DCEDC8"),
        ("Conso moy/mois (u)","B3E5FC"),("Conso min (u)","B3E5FC"),
        ("Conso max (u)","B3E5FC"),("σ mensuelle (u)","B3E5FC"),
        ("σ journalière (u/j)","B3E5FC"),("CV","B3E5FC"),("Conso annuelle (u)","B3E5FC"),
        ("Niveau service","FFF9C4"),("Z","FFF9C4"),
        ("Coût unitaire (TND)","FFE0B2"),("Coût passation (TND)","FFE0B2"),("Taux stockage (%)","FFE0B2"),
        ("Stock sécurité (u)","FFE0B2"),("Coût SS (TND)","FFE0B2"),
        ("EOQ (u)","E8F5E9"),("Nb cmd/an","E8F5E9"),("Point commande (u)","FFE0B2"),
        ("Stock actuel (u)","E8F5E9"),("Commande en cours (u)","E8F5E9"),
    ]
    for col, (label, bg) in enumerate(headers, 1):
        hdr(ws.cell(1, col), label, bg)

    fmts = [
        None,None,None,"0",None,None,None,
        "#,##0.0","#,##0.0","#,##0.0","#,##0.00","#,##0.0000","0.0%","#,##0.0",
        None,"0.00","#,##0.00","#,##0.00","0.0%",
        "#,##0","#,##0.00","#,##0","0.0","#,##0","#,##0","#,##0",
    ]

    for i, row in df_res.iterrows():
        r = i + 2
        vals = [
            row.get("Code article",""),row.get("Description Article",""),
            row.get("Source",""),row.get("LT jours",0),
            row.get("Classe ABC",""),row.get("Classe XYZ",""),row.get("Classe",""),
            row.get("Conso mois moy",0),row.get("Conso min",0),row.get("Conso max",0),
            row.get("Sigma mois",0),row.get("Sigma D (j)",0),row.get("CV",0),row.get("Conso annuelle",0),
            row.get("Niveau service txt",""),row.get("Z val",0),
            row.get("Coût unitaire",0),row.get("Cout passation art",0),row.get("Taux stockage art",0),
            row.get("Stock sécurité",0),row.get("Coût SS",0),
            row.get("EOQ",0),row.get("Nb cmd an",0),row.get("Point de commande",0),
            row.get("Stock actuel",0),row.get("Commande en cours",0),
        ]
        for col, (val, nf) in enumerate(zip(vals, fmts), 1):
            dat(ws.cell(r, col), val, nf)

    abc_fills = {"A": "E8F5E9", "B": "FFFDE7", "C": "F5F5F5"}
    for i, row in df_res.iterrows():
        fill_color = abc_fills.get(row.get("Classe ABC", "C"), "FFFFFF")
        for col in range(1, len(headers) + 1):
            ws.cell(i + 2, col).fill = PatternFill("solid", start_color=fill_color)
            ws.cell(i + 2, col).border = brd
            ws.cell(i + 2, col).font  = Font(size=9, name="Arial")
            ws.cell(i + 2, col).alignment = Alignment(horizontal="center", vertical="center")
        vals = [
            row.get("Code article",""),row.get("Description Article",""),
            row.get("Source",""),row.get("LT jours",0),
            row.get("Classe ABC",""),row.get("Classe XYZ",""),row.get("Classe",""),
            row.get("Conso mois moy",0),row.get("Conso min",0),row.get("Conso max",0),
            row.get("Sigma mois",0),row.get("Sigma D (j)",0),row.get("CV",0),row.get("Conso annuelle",0),
            row.get("Niveau service txt",""),row.get("Z val",0),
            row.get("Coût unitaire",0),row.get("Cout passation art",0),row.get("Taux stockage art",0),
            row.get("Stock sécurité",0),row.get("Coût SS",0),
            row.get("EOQ",0),row.get("Nb cmd an",0),row.get("Point de commande",0),
            row.get("Stock actuel",0),row.get("Commande en cours",0),
        ]
        for col, (val, nf) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(i + 2, col)
            c.value = val
            if nf: c.number_format = nf

    widths = [14,32,10,8,8,8,8,14,12,12,14,14,8,14,14,6,14,16,14,14,14,10,8,16,14,16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 44
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def classify_abc(valeurs_an):
    total = sum(valeurs_an)
    if total == 0: return ["C"] * len(valeurs_an)
    idx_sort = sorted(range(len(valeurs_an)), key=lambda i: valeurs_an[i], reverse=True)
    classes = [""] * len(valeurs_an)
    cumul = 0
    for i in idx_sort:
        cumul += valeurs_an[i] / total
        if cumul <= 0.80:   classes[i] = "A"
        elif cumul <= 0.95: classes[i] = "B"
        else:               classes[i] = "C"
    return classes

def classify_xyz(cv_list):
    return ["X" if cv < 0.20 else "Y" if cv < 0.50 else "Z" for cv in cv_list]

MOIS_ABREV = ["Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]
MOIS_FULL  = ["janvier","février","mars","avril","mai","juin",
               "juillet","août","septembre","octobre","novembre","décembre"]

def detect_month_cols(columns):
    month_cols = []
    for col in columns:
        cl = str(col).lower().strip()
        if cl in [f"m{i}" for i in range(1, 13)]:
            month_cols.append(col)
        elif any(cl.startswith(m.lower()) for m in MOIS_ABREV):
            month_cols.append(col)
        elif any(cl.startswith(m) for m in MOIS_FULL):
            month_cols.append(col)
    return month_cols


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE : ACCUEIL  — délais cards supprimées
# ═══════════════════════════════════════════════════════════════════════════════
def accueil():
    st.markdown("""
    <div style='background:linear-gradient(135deg,#1a3d1a,#2d6a2d);padding:2.2rem 2.5rem;border-radius:16px;
         border:1px solid rgba(245,197,24,.2);margin-bottom:1.5rem;position:relative;overflow:hidden'>
        <div style='position:absolute;top:-30px;right:-30px;width:200px;height:200px;
             background:radial-gradient(circle,rgba(245,197,24,.12),transparent 70%);border-radius:50%'></div>
        <div style='font-size:.7rem;color:#f5c518;letter-spacing:.14em;text-transform:uppercase;margin-bottom:12px;
             background:rgba(245,197,24,.15);display:inline-block;padding:.3rem .8rem;border-radius:20px;
             border:1px solid rgba(245,197,24,.3)'><i class="bi bi-stars"></i> Outil supply chain manager · v3.0</div>
        <h1 style='font-family:"DM Serif Display",serif;font-size:3.3rem;color:#ffffff;margin:.4rem 0;line-height:1.25'>
            Pilotez votre <em style='color:#f5c518'>supply chain</em><br/>avec précision</h1>
        
           </p>
    </div>""", unsafe_allow_html=True)

    # ── Modules disponibles (sans les cartes délais supprimées) ───────────────
    st.markdown("<div style='font-size:1rem;font-weight:700;color:#1a2e1a;margin-bottom:1rem;border-left:3px solid #f5c518;padding-left:10px'>Modules disponibles</div>", unsafe_allow_html=True)
    cols = st.columns(4)
    mods = [
        ("<i class='bi bi-upload bi-xl' style='color:#2d6a2d'></i>", "#2d6a2d", "Calcul automatique",
         "Importez consommations 12 mois → SS · EOQ · ABC/XYZ → Export XLSX"),
        ("<i class='bi bi-sliders bi-xl' style='color:#c9a000'></i>", "#c9a000", "Calculateurs SC",
         "EOQ · SS · Point de réappro. avec courbes"),
        ("<i class='bi bi-bell-fill bi-xl' style='color:#dc3535'></i>", "#dc3535", "Alertes &amp; Stocks",
         "Alertes rupture SS · Déclenchement commandes — utilise les indicateurs du calcul automatique"),
        ("<i class='bi bi-graph-up-arrow bi-xl' style='color:#4a8c3f'></i>", "#4a8c3f", "Évolution &amp; Gains",
         "Suivi SS · saisie par code · évaluation économies réalisées"),
    ]
    for col, (ic, accent, ti, de) in zip(cols, mods):
        with col:
            st.markdown(
                f"<div style='background:#ffffff;border:1px solid #d4e0d4;border-radius:12px;padding:1.3rem;"
                f"border-top:3px solid {accent};box-shadow:0 2px 8px rgba(26,61,26,.05)'>"
                f"<div style='font-size:1.6rem;margin-bottom:10px'>{ic}</div>"
                f"<div style='font-size:.9rem;font-weight:700;color:#1a2e1a;margin-bottom:6px'>{ti}</div>"
                f"<div style='font-size:.78rem;color:#7a9a7a;line-height:1.55'>{de}</div></div>",
                unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE : CALCUL AUTOMATIQUE
# ═══════════════════════════════════════════════════════════════════════════════
def import_calcul():
    st.markdown('<h2><i class="bi bi-cloud-upload" style="color:#2d6a2d;margin-right:10px"></i>Calcul automatique — Stock de sécurité &amp; EOQ</h2>', unsafe_allow_html=True)
    st.markdown(
        "<p style='color:#7a9a7a'>Importez votre tableau de consommations 12 mois — "
        "σD, demande moyenne, classification ABC/XYZ et indicateurs SC calculés automatiquement "
        "avec le <strong>niveau de service adapté à chaque classe</strong> et les <strong>coûts propres à chaque article</strong>.</p>",
        unsafe_allow_html=True)

    st.markdown("""
<div style='display:grid;grid-template-columns:repeat(3,1fr);gap:.8rem;margin-bottom:1rem'>
  <div style='background:#e8f5e9;border:1px solid #c8e6c9;border-radius:10px;padding:.8rem 1rem;border-left:4px solid #2d6a2d'>
    <div style='font-size:.65rem;color:#2d6a2d;text-transform:uppercase;font-weight:700;margin-bottom:4px'>Classe A — Articles stratégiques</div>
    <div style='font-family:"DM Mono",monospace;font-size:1.1rem;color:#1a3d1a;font-weight:700'>Z = 2.33 → NS 99%</div>
    <div style='font-size:.72rem;color:#4a8c3f;margin-top:3px'>Top 80% de la valeur annuelle</div>
  </div>
  <div style='background:#fffde7;border:1px solid #fff9c4;border-radius:10px;padding:.8rem 1rem;border-left:4px solid #c9a000'>
    <div style='font-size:.65rem;color:#c9a000;text-transform:uppercase;font-weight:700;margin-bottom:4px'>Classe B — Articles intermédiaires</div>
    <div style='font-family:"DM Mono",monospace;font-size:1.1rem;color:#7a5c00;font-weight:700'>Z = 1.65 → NS 95%</div>
    <div style='font-size:.72rem;color:#c9a000;margin-top:3px'>80–95% de la valeur annuelle</div>
  </div>
  <div style='background:#f5f5f5;border:1px solid #e0e0e0;border-radius:10px;padding:.8rem 1rem;border-left:4px solid #9e9e9e'>
    <div style='font-size:.65rem;color:#757575;text-transform:uppercase;font-weight:700;margin-bottom:4px'>Classe C — Articles courants</div>
    <div style='font-family:"DM Mono",monospace;font-size:1.1rem;color:#424242;font-weight:700'>Z = 1.28 → NS 90%</div>
    <div style='font-size:.72rem;color:#9e9e9e;margin-top:3px'>95–100% de la valeur annuelle</div>
  </div>
</div>
""", unsafe_allow_html=True)

    with st.expander("Format attendu du fichier Excel"):
        st.markdown("""
**Le fichier peut avoir 1 ou 2 feuilles — l'application traite chaque feuille séparément.**

| Code article | Description | Source | Coût unitaire | Coût passation | Taux stockage | Jan | Fév | … | Déc | Stock actuel | Commande en cours |
|---|---|---|---|---|---|---|---|---|---|---|---|
| 40007 | NYSOSEL | Export | 59.23 | 180 | 0.22 | 310 | 324 | … | 318 | 1200 | 0 |

- **`Coût passation`** et **`Taux stockage`** sont propres à chaque article
- Les colonnes mois : **Jan/Fév…**, **M1/M2…** ou **Janvier/Février…**
- `Source` : **Export** (4 mois) · **Local** (2 sem.) · **BM** (3 sem.)
- Le **niveau de service** est attribué automatiquement 
""")

    st.divider()
    st.markdown('<h4><i class="bi bi-folder2-open" style="color:#2d6a2d;margin-right:8px"></i>Chargement du fichier</h4>', unsafe_allow_html=True)
    uploaded = st.file_uploader("Déposez votre fichier Excel (.xlsx / .xls)",
                                 type=["xlsx", "xls"], key="ic_upload")
    use_demo = st.checkbox("Utiliser les données de démo", value=not bool(uploaded), key="ic_demo")

    all_sheets_data = {}

    if uploaded:
        try:
            all_raw     = pd.read_excel(uploaded, sheet_name=None, header=None)
            sheet_names = list(all_raw.keys())
            if len(sheet_names) > 1:
                st.info(f"**{len(sheet_names)} feuilles détectées :** {', '.join(sheet_names)}")
                feuille_active    = st.radio("Feuille à analyser", sheet_names, horizontal=True, key="ic_sheet_sel")
                sheets_to_process = [feuille_active]
            else:
                sheets_to_process = sheet_names

            for sheet_sel in sheets_to_process:
                raw  = all_raw[sheet_sel]
                hrow = 0
                for i, row in raw.iterrows():
                    row_str = " ".join([str(v).lower() for v in row if pd.notna(v)])
                    if any(k in row_str for k in ["code", "article", "jan", "m1", "source"]):
                        hrow = i; break

                df_raw = pd.read_excel(uploaded, sheet_name=sheet_sel, header=hrow)
                df_raw.columns = [str(c).strip() for c in df_raw.columns]
                df_raw = df_raw.dropna(how="all")

                cmap = {}
                for col in df_raw.columns:
                    cl = col.lower().strip()
                    if "code" in cl and "article" in cl:                                  cmap["Code article"] = col
                    elif "description" in cl or "désignation" in cl:                      cmap["Description Article"] = col
                    elif cl == "source":                                                   cmap["Source"] = col
                    elif "coût unitaire" in cl or "cout unitaire" in cl or cl == "prix":  cmap["Coût unitaire"] = col
                    elif "passation" in cl or "cout commande" in cl or "coût commande" in cl: cmap["Coût passation"] = col
                    elif "taux" in cl and ("stock" in cl or "stockage" in cl):             cmap["Taux stockage"] = col
                    elif "stock actuel" in cl or ("stock" in cl and "actuel" in cl):       cmap["Stock actuel"] = col
                    elif "commande en cours" in cl or "cmd en cours" in cl:                cmap["Commande en cours"] = col

                df_s = df_raw.rename(columns={v: k for k, v in cmap.items()})
                month_cols = detect_month_cols(df_s.columns)

                for c in ["Coût unitaire","Coût passation","Taux stockage",
                          "Stock actuel","Commande en cours"] + month_cols:
                    if c in df_s.columns:
                        df_s[c] = pd.to_numeric(df_s[c], errors="coerce").fillna(0)

                if "Code article" in df_s.columns:
                    df_s = df_s[df_s["Code article"].notna()]
                    df_s = df_s[df_s["Code article"].astype(str).str.strip() != ""]
                if "Source" not in df_s.columns:
                    df_s["Source"] = "Export"

                df_s.attrs["month_cols"] = month_cols
                all_sheets_data[sheet_sel] = df_s
                st.success(f"Feuille **{sheet_sel}** : {len(df_s)} articles · "
                           f"{len(month_cols)} mois ({', '.join(month_cols[:3])}{'…' if len(month_cols)>3 else ''})")

        except Exception as e:
            st.error(f"Erreur : {e}")

    if use_demo and not all_sheets_data:
        rng  = np.random.default_rng(42)
        base = [324, 5886, 509, 32460, 959, 3747, 3147, 4432, 16, 22]
        md   = {f"M{i+1}": [int(b*(0.85+0.30*rng.random())) for b in base] for i in range(12)}
        demo = pd.DataFrame({
            "Code article":       [40007,40010,40011,40014,40015,40036,40042,40055,40062,40063],
            "Description Article":["NYSOSEL","TERRE TONSIL","PAPIER FILTRE","TOILE FILTRANTE",
                                   "PERLITE PAF1","SEL MARIN FIN","MYVEROL 18-04","TRISYL",
                                   "VITAMINE A","VITAMINE E"],
            "Source":             ["Export","Export","Local","Export","BM","Local","Export","BM","Export","Local"],
            "Coût unitaire":      [59.23,1.709,3.823,49.787,1.700,0.85,2.10,8.189,12.5,18.0],
            "Coût passation":     [200,150,100,250,120,80,130,160,90,110],
            "Taux stockage":      [0.22,0.18,0.20,0.25,0.18,0.15,0.20,0.22,0.20,0.18],
            "Stock actuel":       [1200,8000,500,50000,2000,5000,8000,9000,30,50],
            "Commande en cours":  [0,5000,0,0,1000,0,3000,0,0,0],
            **md
        })
        demo.attrs["month_cols"] = [f"M{i+1}" for i in range(12)]
        all_sheets_data["Démo"] = demo
        st.info("Données de démo chargées")

    if not all_sheets_data:
        return

    for sheet_name, df_source in all_sheets_data.items():
        if len(all_sheets_data) > 1:
            st.markdown(f'---\n<h3><i class="bi bi-file-earmark-spreadsheet" style="color:#2d6a2d"></i> Feuille : <strong>{sheet_name}</strong></h3>', unsafe_allow_html=True)

        month_cols = df_source.attrs.get("month_cols", [])

        with st.expander(f"Tableau importé — {sheet_name}", expanded=True):
            disp_cols = (["Code article","Description Article","Source","Coût unitaire",
                          "Coût passation","Taux stockage"]
                         + month_cols
                         + [c for c in ["Stock actuel","Commande en cours"] if c in df_source.columns])
            st.dataframe(df_source[[c for c in disp_cols if c in df_source.columns]],
                         use_container_width=True, hide_index=True)

        need_cu = "Coût unitaire"  not in df_source.columns or df_source["Coût unitaire"].sum()  == 0
        need_sc = "Coût passation" not in df_source.columns or df_source["Coût passation"].sum() == 0
        need_tx = "Taux stockage"  not in df_source.columns or df_source["Taux stockage"].sum()  == 0

        if need_cu or need_sc or need_tx:
            missing = []
            if need_cu: missing.append("Coût unitaire")
            if need_sc: missing.append("Coût passation")
            if need_tx: missing.append("Taux stockage (%)")
            st.warning(f"Colonnes absentes ou nulles : **{', '.join(missing)}** — saisissez les valeurs ci-dessous :")
            saisie_data = []
            for _, row in df_source.iterrows():
                saisie_data.append({
                    "Code article":   str(row.get("Code article", "")),
                    "Description":    str(row.get("Description Article", ""))[:30],
                    "Coût unitaire":  float(row.get("Coût unitaire", 0) or 0)  if not need_cu else 0.0,
                    "Coût passation": float(row.get("Coût passation", 150) or 150) if not need_sc else 150.0,
                    "Taux stockage":  float(row.get("Taux stockage", 0.20) or 0.20) if not need_tx else 0.20,
                })
            df_saisie = pd.DataFrame(saisie_data)
            edited = st.data_editor(df_saisie, use_container_width=True, hide_index=True, key=f"saisie_{sheet_name}",
                column_config={
                    "Code article":   st.column_config.TextColumn("Code", disabled=True),
                    "Description":    st.column_config.TextColumn("Description", disabled=True),
                    "Coût unitaire":  st.column_config.NumberColumn("Coût unitaire (TND)", min_value=0.0, step=0.01, format="%.4f"),
                    "Coût passation": st.column_config.NumberColumn("Coût passation (TND)", min_value=0.0, step=1.0, format="%.1f"),
                    "Taux stockage":  st.column_config.NumberColumn("Taux stockage (ex: 0.20)", min_value=0.0, max_value=1.0, step=0.01, format="%.2f"),
                })
            df_source = df_source.copy()
            code_map_cu = dict(zip(edited["Code article"].astype(str), edited["Coût unitaire"]))
            code_map_sc = dict(zip(edited["Code article"].astype(str), edited["Coût passation"]))
            code_map_tx = dict(zip(edited["Code article"].astype(str), edited["Taux stockage"]))
            if need_cu: df_source["Coût unitaire"]  = df_source["Code article"].astype(str).map(code_map_cu).fillna(0)
            if need_sc: df_source["Coût passation"] = df_source["Code article"].astype(str).map(code_map_sc).fillna(150)
            if need_tx: df_source["Taux stockage"]  = df_source["Code article"].astype(str).map(code_map_tx).fillna(0.20)

        pre_vals_an = []
        for _, row in df_source.iterrows():
            cm = [float(row.get(mc, 0) or 0) for mc in month_cols] if month_cols else [float(row.get("Consommation", 0) or 0)]
            cu = float(row.get("Coût unitaire", 0) or 0)
            pre_vals_an.append(np.mean(cm) * 12 * cu)
        pre_abc = classify_abc(pre_vals_an)

        results = []
        for i, (_, row) in enumerate(df_source.iterrows()):
            source = str(row.get("Source", "Export")).strip()
            cu     = float(row.get("Coût unitaire", 0)    or 0)
            sc_art = float(row.get("Coût passation", 150) or 150)
            tx_art = float(row.get("Taux stockage", 0.20) or 0.20)
            stk    = float(row.get("Stock actuel", 0)     or 0)
            cmd    = float(row.get("Commande en cours", 0) or 0)
            lt_j   = get_lt_jours(source)

            conso_mois = ([float(row.get(mc, 0) or 0) for mc in month_cols]
                          if month_cols else [float(row.get("Consommation", 0) or 0)])
            d_moy    = np.mean(conso_mois)
            d_moy_j  = d_moy / 30
            sigma_m  = float(np.std(conso_mois, ddof=1)) if len(conso_mois) > 1 else d_moy * 0.15
            sigma_j  = sigma_m / 30
            cv       = sigma_m / d_moy if d_moy > 0 else 0
            conso_an = d_moy * 12
            val_an   = conso_an * cu

            abc_prov = pre_abc[i]
            Z_art    = Z_PAR_CLASSE[abc_prov]
            ns_txt   = NS_PAR_CLASSE[abc_prov]

            ss = calc_ss(sigma_j, lt_j, Z_art)
            pr = calc_pr(d_moy_j, lt_j, ss)
            eoq, nb_cmd, _ = calc_eoq(conso_an, cu, sc_cost=sc_art, taux=tx_art)
            css = round(ss * cu, 2) if cu > 0 else 0

            results.append({
                "Code article":        row.get("Code article", ""),
                "Description Article": row.get("Description Article", ""),
                "Source":              source,
                "LT jours":            lt_j,
                "Conso mois moy":      round(d_moy, 1),
                "Conso min":           round(min(conso_mois), 1),
                "Conso max":           round(max(conso_mois), 1),
                "Sigma mois":          round(sigma_m, 2),
                "Sigma D (j)":         round(sigma_j, 4),
                "CV":                  round(cv, 3),
                "D jour":              round(d_moy_j, 4),
                "Conso annuelle":      round(conso_an, 1),
                "Valeur annuelle":     round(val_an, 2),
                "Z val":               Z_art,
                "Niveau service txt":  ns_txt,
                "Coût unitaire":       cu,
                "Cout passation art":  sc_art,
                "Taux stockage art":   tx_art,
                "Stock sécurité":      ss,
                "Coût SS":             css,
                "EOQ":                 eoq,
                "Nb cmd an":           nb_cmd,
                "Point de commande":   pr,
                "Stock actuel":        stk,
                "Commande en cours":   cmd,
                "_conso_mois":         conso_mois,
            })

        df_res = pd.DataFrame(results)
        df_res["Classe ABC"] = classify_abc(df_res["Valeur annuelle"].tolist())
        df_res["Classe XYZ"] = classify_xyz(df_res["CV"].tolist())
        df_res["Classe"]     = df_res["Classe ABC"] + df_res["Classe XYZ"]
        df_res["Z val"]             = df_res["Classe ABC"].map(Z_PAR_CLASSE)
        df_res["Niveau service txt"] = df_res["Classe ABC"].map(NS_PAR_CLASSE)
        df_res["Stock sécurité"]    = df_res.apply(lambda r: calc_ss(r["Sigma D (j)"], r["LT jours"], r["Z val"]), axis=1)
        df_res["Coût SS"]           = (df_res["Stock sécurité"] * df_res["Coût unitaire"]).round(2)
        df_res["Point de commande"] = df_res.apply(lambda r: calc_pr(r["D jour"], r["LT jours"], r["Stock sécurité"]), axis=1)

        # ── Section 1 — Analyse ───────────────────────────────────────────────
        st.divider()
        st.markdown('<h3><i class="bi bi-table" style="color:#2d6a2d;margin-right:8px"></i>Analyse des consommations</h3>', unsafe_allow_html=True)

        df_stats = df_res[["Code article","Description Article","Source",
            "Conso mois moy","Conso min","Conso max","Sigma mois","CV",
            "Classe ABC","Classe XYZ","Classe","Niveau service txt","Z val"]].copy()

        st.dataframe(df_stats, use_container_width=True, hide_index=True,
            column_config={
                "Conso mois moy":     st.column_config.NumberColumn("Moy./mois (u)", format="%.1f"),
                "Conso min":          st.column_config.NumberColumn("Min (u)", format="%.1f"),
                "Conso max":          st.column_config.NumberColumn("Max (u)", format="%.1f"),
                "Sigma mois":         st.column_config.NumberColumn("σ mensuel (u)", format="%.2f"),
                "CV":                 st.column_config.NumberColumn("CV", format="%.1%%"),
                "Classe ABC":         st.column_config.TextColumn("ABC"),
                "Classe XYZ":         st.column_config.TextColumn("XYZ"),
                "Classe":             st.column_config.TextColumn("Classe"),
                "Niveau service txt": st.column_config.TextColumn("Niveau service"),
                "Z val":              st.column_config.NumberColumn("Z", format="%.2f"),
            })

        st.markdown("""
<div style='display:grid;grid-template-columns:1fr 1fr;gap:.8rem;margin:.6rem 0 1rem'>
  <div style='background:#ffffff;border:1px solid #d4e0d4;border-radius:10px;padding:.9rem 1.1rem'>
    <div style='font-size:.68rem;color:#7a9a7a;text-transform:uppercase;font-weight:600;margin-bottom:.4rem'>ABC — Valeur annuelle cumulée</div>
    <div style='font-size:.8rem;line-height:2'>
      <span style='background:#2d6a2d;color:#fff;padding:2px 8px;border-radius:4px;font-weight:700'>A</span>&nbsp; 80% valeur → <strong>NS 99%</strong> (Z=2.33)&nbsp;&nbsp;
      <span style='background:#c9a000;color:#fff;padding:2px 8px;border-radius:4px;font-weight:700'>B</span>&nbsp; 80–95% → <strong>NS 95%</strong> (Z=1.65)&nbsp;&nbsp;
      <span style='background:#9e9e9e;color:#fff;padding:2px 8px;border-radius:4px;font-weight:700'>C</span>&nbsp; reste → <strong>NS 90%</strong> (Z=1.28)
    </div>
  </div>
  <div style='background:#ffffff;border:1px solid #d4e0d4;border-radius:10px;padding:.9rem 1.1rem'>
    <div style='font-size:.68rem;color:#7a9a7a;text-transform:uppercase;font-weight:600;margin-bottom:.4rem'>XYZ — Coefficient de variation</div>
    <div style='font-size:.8rem;line-height:2'>
      <span style='background:#1565c0;color:#fff;padding:2px 8px;border-radius:4px;font-weight:700'>X</span>&nbsp; CV&lt;20% stable&nbsp;&nbsp;
      <span style='background:#880e4f;color:#fff;padding:2px 8px;border-radius:4px;font-weight:700'>Y</span>&nbsp; CV 20–50% variable&nbsp;&nbsp;
      <span style='background:#e65100;color:#fff;padding:2px 8px;border-radius:4px;font-weight:700'>Z</span>&nbsp; CV&gt;50% irrégulier
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── Section 2 — Détail article ────────────────────────────────────────
        st.divider()
        st.markdown('<h3><i class="bi bi-search" style="color:#2d6a2d;margin-right:8px"></i>Détail par article</h3>', unsafe_allow_html=True)

        art_labels = [f"{row['Code article']} — {row['Description Article']} [{row['Classe']}]"
                      for _, row in df_res.iterrows()]
        sel_label  = st.selectbox("Sélectionner un article", art_labels, key=f"sel_art_{sheet_name}")
        sel_idx    = art_labels.index(sel_label)
        art        = df_res.iloc[sel_idx]
        conso_mois_art = art["_conso_mois"]
        n_mois_art     = len(conso_mois_art)
        mois_x         = MOIS_ABREV[:n_mois_art] if n_mois_art <= 12 else [f"M{i+1}" for i in range(n_mois_art)]

        abc_col_map = {"A": C_GREEN, "B": C_GOLD, "C": C_MUTED}
        xyz_col_map = {"X": "#1565c0", "Y": "#880e4f", "Z": "#e65100"}
        abc_col = abc_col_map.get(art["Classe ABC"], C_GREEN)
        xyz_col = xyz_col_map.get(art["Classe XYZ"], C_MUTED)

        st.markdown(
            f"<div style='background:#ffffff;border:1px solid #d4e0d4;border-radius:12px;"
            f"padding:1rem 1.4rem;margin-bottom:1rem;border-left:4px solid {abc_col}'>"
            f"<span style='font-size:1.1rem;font-weight:700;color:#1a2e1a'>{art['Code article']} — {art['Description Article']}</span>"
            f"&nbsp;&nbsp;"
            f"<span style='background:{abc_col};color:#fff;padding:3px 10px;border-radius:20px;font-size:.8rem;font-weight:700'>{art['Classe ABC']}</span>"
            f"&nbsp;<span style='background:{xyz_col};color:#fff;padding:3px 10px;border-radius:20px;font-size:.8rem;font-weight:700'>{art['Classe XYZ']}</span>"
            f"&nbsp;&nbsp;<span style='font-size:.82rem;color:#7a9a7a'>"
            f"Source : {art['Source']} · Délai : {art['LT jours']}j · NS : <strong>{art['Niveau service txt']}</strong> (Z={art['Z val']})</span></div>",
            unsafe_allow_html=True)

        c1,c2,c3,c4,c5 = st.columns(5)
        with c1: st.markdown(mcard("Demande moy./mois", fmt(art["Conso mois moy"],1), "u/mois", color=C_GREEN,  icon="bi-graph-down"), unsafe_allow_html=True)
        with c2: st.markdown(mcard("Demande moy./jour", fmt(art["D jour"],3),         "u/j",    color=C_GREEN2, icon="bi-activity"),   unsafe_allow_html=True)
        with c3: st.markdown(mcard("σ mensuelle",       fmt(art["Sigma mois"],2),     "u/mois", color=C_PURPLE, icon="bi-distribute-vertical"), unsafe_allow_html=True)
        with c4: st.markdown(mcard("σ journalière",     fmt(art["Sigma D (j)"],4),    "u/j",    color=C_PURPLE, icon="bi-distribute-vertical"), unsafe_allow_html=True)
        with c5: st.markdown(mcard("Coeff. variation",  f"{art['CV']*100:.1f}%",      "CV",     color=xyz_col,  icon="bi-percent"),    unsafe_allow_html=True)
        st.markdown("<div style='margin-bottom:.8rem'></div>", unsafe_allow_html=True)

        c1,c2,c3,c4,c5 = st.columns(5)
        with c1: st.markdown(mcard("Coût unitaire",    fmt(art["Coût unitaire"],4),     "TND/u",  color=C_MUTED,  icon="bi-tag"),           unsafe_allow_html=True)
        with c2: st.markdown(mcard("Coût passation",   fmt(art["Cout passation art"],1),"TND/cmd",color=C_MUTED,  icon="bi-receipt"),        unsafe_allow_html=True)
        with c3: st.markdown(mcard("Taux stockage",    f"{art['Taux stockage art']*100:.1f}%", "%/an", color=C_MUTED, icon="bi-percent"),    unsafe_allow_html=True)
        with c4: st.markdown(mcard("Niveau de service",art["Niveau service txt"],       f"Z={art['Z val']}", color=abc_col, icon="bi-shield-check"), unsafe_allow_html=True)
        with c5: st.markdown(mcard("Valeur annuelle",  f"{art['Valeur annuelle']:,.0f}","TND/an", color=C_GREEN2, icon="bi-cash-stack"),     unsafe_allow_html=True)
        st.markdown("<div style='margin-bottom:.8rem'></div>", unsafe_allow_html=True)

        c1,c2,c3,c4,c5 = st.columns(5)
        with c1: st.markdown(mcard("Stock de sécurité",fmtInt(art["Stock sécurité"]),  "unités",  color=C_GOLD,   icon="bi-shield-fill-check"), unsafe_allow_html=True)
        with c2: st.markdown(mcard("Point de commande",fmtInt(art["Point de commande"]),"unités",  color=C_RED,    icon="bi-flag-fill"),          unsafe_allow_html=True)
        with c3: st.markdown(mcard("EOQ",              fmtInt(art["EOQ"]),             "u/cmd",   color=C_GREEN,  icon="bi-arrow-repeat"),       unsafe_allow_html=True)
        with c4: st.markdown(mcard("Coût SS",          f"{art['Coût SS']:,.2f}",       "TND",     color=C_GOLD,   icon="bi-wallet2"),            unsafe_allow_html=True)
        with c5: st.markdown(mcard("Nb cmd/an",        fmt(art["Nb cmd an"],1),        "cmd/an",  color=C_GREEN2, icon="bi-cart"),               unsafe_allow_html=True)
        st.markdown("<div style='margin-bottom:.8rem'></div>", unsafe_allow_html=True)

        st.markdown(
            fbox(f"SS calculé — {art['Code article']}",
                 f"SS = Z × σD × √L = {art['Z val']} × {art['Sigma D (j)']} × √{art['LT jours']}",
                 f"= {art['Z val']} × {art['Sigma D (j)']} × {round(math.sqrt(art['LT jours']),3)}"
                 f" = <strong>{art['Stock sécurité']}</strong> unités "
                 f"(Niveau de service {art['Niveau service txt']} — Classe {art['Classe ABC']})"),
            unsafe_allow_html=True)

        fig_art = go.Figure()
        fig_art.add_trace(go.Bar(x=mois_x, y=conso_mois_art, name="Consommation",
            marker_color=[C_RED if c < art["Conso mois moy"] - art["Sigma mois"]
                          else C_GOLD if c > art["Conso mois moy"] + art["Sigma mois"]
                          else C_GREEN for c in conso_mois_art]))
        fig_art.add_hline(y=art["Conso mois moy"], line_dash="dash", line_color=C_GREEN2, line_width=2,
                           annotation_text=f"μ = {art['Conso mois moy']:.1f}", annotation_font_color=C_GREEN2)
        fig_art.add_hline(y=art["Conso mois moy"] + art["Sigma mois"], line_dash="dot", line_color=C_PURPLE, line_width=1.5,
                           annotation_text=f"μ+σ = {art['Conso mois moy']+art['Sigma mois']:.1f}", annotation_font_color=C_PURPLE)
        fig_art.add_hline(y=max(0, art["Conso mois moy"] - art["Sigma mois"]), line_dash="dot", line_color=C_PURPLE, line_width=1.5,
                           annotation_text=f"μ−σ = {max(0,art['Conso mois moy']-art['Sigma mois']):.1f}", annotation_font_color=C_PURPLE)
        fig_art.update_layout(**light_layout(), height=300,
            xaxis=dict(title="Mois", gridcolor="rgba(0,0,0,.06)"),
            yaxis=dict(title="Consommation (u)", gridcolor="rgba(0,0,0,.06)"),
            title=f"Historique 12 mois — {art['Code article']} · {art['Description Article']}")
        st.plotly_chart(fig_art, use_container_width=True)

        # ── Section 3 — Tableau SC ────────────────────────────────────────────
        st.divider()
        st.markdown('<h3><i class="bi bi-grid-3x3-gap-fill" style="color:#2d6a2d;margin-right:8px"></i>Indicateurs SC — tous articles</h3>', unsafe_allow_html=True)

        disp_sc = df_res[["Code article","Description Article","Source","Classe",
            "Niveau service txt","Z val","LT jours","Conso mois moy","Sigma D (j)","CV",
            "Cout passation art","Taux stockage art",
            "Stock sécurité","Coût SS","EOQ","Point de commande","Stock actuel","Commande en cours"]].copy()

        st.dataframe(disp_sc, use_container_width=True, hide_index=True,
            column_config={
                "Niveau service txt":  st.column_config.TextColumn("NS"),
                "Z val":               st.column_config.NumberColumn("Z", format="%.2f"),
                "LT jours":            st.column_config.NumberColumn("Délai (j)", format="%d"),
                "Conso mois moy":      st.column_config.NumberColumn("Conso moy./mois", format="%.1f"),
                "Sigma D (j)":         st.column_config.NumberColumn("σD (u/j)", format="%.4f"),
                "CV":                  st.column_config.NumberColumn("CV", format="%.1%%"),
                "Cout passation art":  st.column_config.NumberColumn("Coût passation (TND)", format="%.1f"),
                "Taux stockage art":   st.column_config.NumberColumn("Taux stockage", format="%.1%%"),
                "Stock sécurité":      st.column_config.NumberColumn("SS (u)", format="%d"),
                "Coût SS":             st.column_config.NumberColumn("Coût SS (TND)", format="%.2f"),
                "EOQ":                 st.column_config.NumberColumn("EOQ (u)", format="%d"),
                "Point de commande":   st.column_config.NumberColumn("Point commande (u)", format="%d"),
                "Stock actuel":        st.column_config.NumberColumn("Stock actuel (u)", format="%d"),
                "Commande en cours":   st.column_config.NumberColumn("Cmd en cours (u)", format="%d"),
            })

        

        st.divider()
        c1,c2,c3,c4 = st.columns(4)
        n_A = (df_res["Classe ABC"]=="A").sum()
        with c1: st.markdown(mcard("Articles traités",  str(len(df_res)), "articles",         color=C_GREEN,  icon="bi-boxes"),            unsafe_allow_html=True)
        with c2: st.markdown(mcard("Coût total SS",     f"{df_res['Coût SS'].sum():,.0f}", "TND immobilisés", color=C_GOLD, icon="bi-currency-exchange"), unsafe_allow_html=True)
       

        # ── Section 4 — Export ────────────────────────────────────────────────
        st.divider()
        st.markdown('<h3><i class="bi bi-file-earmark-excel-fill" style="color:#217346;margin-right:8px"></i>Export XLSX — Indicateurs SC</h3>', unsafe_allow_html=True)
        st.info("Le fichier contient les indicateurs calculés avec classification ABC/XYZ, niveaux de service automatiques et coûts propres à chaque article.")

        df_exp = df_res.drop(columns=["_conso_mois","Valeur annuelle"], errors="ignore")
        buf    = build_excel_indicateurs(df_exp)
        st.download_button(
            label="Télécharger le fichier Excel",
            data=buf,
            file_name=f"medoil_indicateurs_{sheet_name}_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

        # ── Sauvegarde session pour Alertes ───────────────────────────────────
        if "df_resultats" not in st.session_state:
            st.session_state["df_resultats"] = {}
        st.session_state["df_resultats"][sheet_name] = df_exp


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE : CALCULATEURS SC  — onglet KPIs stock supprimé
# ═══════════════════════════════════════════════════════════════════════════════
def calculateurs():
    st.markdown('<h2><i class="bi bi-sliders" style="color:#2d6a2d;margin-right:10px"></i>Calculateurs Supply Chain</h2>', unsafe_allow_html=True)

    # Seulement 3 onglets — KPIs stock supprimé
    tab_ss, tab_eoq, tab_rp = st.tabs(["Stock de sécurité", "EOQ (Wilson)", "Point de réappro."])

    with tab_ss:
        st.markdown(fbox("Formule", "SS = Z × σD × √L",
                         "Z = facteur service · σD = écart-type demande (u/j) · L = délai (j) · Délais : Export=120j · Local=14j · BM=3j",
                         icon="bi-shield-fill-check"), unsafe_allow_html=True)

        c1, c2, c3 = st.columns(3)
        with c1:
            source_ss = st.selectbox("Source approvisionnement", ["Export", "Local", "BM"], key="ss_source")
            lt_auto   = get_lt_jours(source_ss)
            st.info(f"Délai  : **{lt_auto} jours**")
            sd  = st.number_input("Demande moy. (u/j)",          value=50.0, step=1.0, key="ss_sd")
            ss2 = st.number_input("Écart-type demande σD (u/j)", value=8.0,  step=0.5, key="ss_sigma")
        with c2:
            scu = st.number_input("Coût unitaire (TND)", value=25.0, step=1.0, key="ss_cu")
        with c3:
            Z_MAP_SS = {"95%": 1.65, "96%": 1.75, "97%": 1.88, "98%": 2.05, "99%": 2.33,
                        "90%": 1.28, "85%": 1.04, "80%": 0.84}
            szo = st.selectbox("Niveau de service", list(Z_MAP_SS.keys()), index=0, key="ss_ns")

        Z2      = Z_MAP_SS[szo]
        SS_calc = Z2 * ss2 * math.sqrt(lt_auto)

        c = st.columns(3)
        for i, (l, v, u, col) in enumerate([
            ("SS calculé",         fmtInt(SS_calc),       "unités", C_GREEN),
            ("Z × σD × √L",        f"{Z2} × {ss2} × √{lt_auto}", "détail calcul", C_GREEN2),
            ("Coût immobilisé SS", fmtInt(SS_calc * scu), "TND",    C_GOLD),
        ]):
            with c[i]: st.markdown(mcard(l, v, u, color=col), unsafe_allow_html=True)

        st.markdown("<div style='margin-bottom:14px'></div>", unsafe_allow_html=True)
        z_vals  = [0.84, 1.04, 1.28, 1.65, 1.88, 2.05, 2.33]
        ns_vals = [80,   85,   90,   95,   97,   98,   99]
        ss_v    = [round(z * ss2 * math.sqrt(lt_auto)) for z in z_vals]
        current_pct = int(szo.replace("%", ""))
        fig = go.Figure(go.Bar(
            x=[f"{n}%" for n in ns_vals], y=ss_v,
            marker_color=[C_GOLD if n == current_pct else C_GREEN for n in ns_vals],
            text=ss_v, textposition="outside"))
        fig.update_layout(**light_layout(), height=260,
            xaxis=dict(title="Niveau de service", gridcolor="rgba(0,0,0,.06)"),
            yaxis=dict(title="SS (unités)", gridcolor="rgba(0,0,0,.06)"))
        st.plotly_chart(fig, use_container_width=True)

    with tab_eoq:
        st.markdown(fbox("Modèle de Wilson", "EOQ = √(2×D×Sc / (Sh×Cu))",
                         "D=demande annuelle · Sc=coût passation · Sh=taux stockage · Cu=coût unitaire",
                         icon="bi-calculator-fill"), unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        with c1:
            eD  = st.number_input("Demande annuelle (u)", value=12000, step=100,  key="eoq_D")
            eSc = st.number_input("Coût passation (TND)", value=150.0, step=5.0,  key="eoq_Sc")
        with c2:
            eCu = st.number_input("Coût unitaire (TND)",  value=80.0,  step=1.0,  key="eoq_Cu")
            eSh = st.number_input("Taux stockage (%)",    value=20.0,  step=1.0,  key="eoq_Sh")
        with c3:
            eLT = st.number_input("Délai fourni. (j)",    value=10,    step=1,    key="eoq_LT")
            eDy = st.number_input("Jours ouvr./an",       value=250,   step=1,    key="eoq_Dy")

        h   = (eSh / 100) * eCu
        EOQ = math.sqrt(2 * eD * eSc / h) if h > 0 else 0
        nb  = eD / EOQ if EOQ > 0 else 0
        T   = eDy / nb if nb > 0 else 0
        Cp  = nb * eSc
        Cs  = (EOQ / 2) * h

        st.markdown("---")
        c = st.columns(3)
        data = [("EOQ",               fmtInt(EOQ),     "u/cmd",  C_GREEN),
                ("Nb cmd/an",         fmt(nb, 1),      "cmds",   C_MUTED),
                ("Périodicité",       fmtInt(T),       "jours",  C_MUTED),
                ("Coût passation/an", fmtInt(Cp),      "TND",    C_GOLD),
                ("Coût stockage/an",  fmtInt(Cs),      "TND",    C_GOLD),
                ("Coût total min.",   fmtInt(Cp + Cs), "TND",    C_GREEN2)]
        for i, (l, v, u, col) in enumerate(data):
            with c[i % 3]:
                st.markdown(mcard(l, v, u, color=col), unsafe_allow_html=True)
                st.markdown("<div style='margin-bottom:10px'></div>", unsafe_allow_html=True)

        qty_r = np.linspace(max(10, EOQ * .2), EOQ * 2.5, 200)
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=qty_r, y=[(eD / q) * eSc for q in qty_r], name="Passation", line=dict(color=C_GREEN, width=2)))
        fig.add_trace(go.Scatter(x=qty_r, y=[(q / 2) * h    for q in qty_r], name="Stockage",  line=dict(color=C_GOLD, width=2)))
        fig.add_trace(go.Scatter(x=qty_r, y=[(eD / q) * eSc + (q / 2) * h for q in qty_r], name="Total", line=dict(color=C_GREEN2, width=2.5)))
        fig.add_vline(x=EOQ, line_dash="dash", line_color=C_RED,
                      annotation_text=f"EOQ={fmtInt(EOQ)}", annotation_font_color=C_RED)
        fig.update_layout(**light_layout(), height=300,
            xaxis=dict(title="Quantité (u)", gridcolor="rgba(0,0,0,.06)"),
            yaxis=dict(title="Coût/an (TND)", gridcolor="rgba(0,0,0,.06)"))
        st.plotly_chart(fig, use_container_width=True)

    with tab_rp:
        st.markdown(fbox("Formule", "PR = (Dmoy × LT) + SS",
                         "Dmoy = demande/jour · LT = délai (j) · SS = stock de sécurité",
                         icon="bi-flag-fill"), unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        with c1:
            src_rp = st.selectbox("Source", ["Export", "Local", "BM"], key="rp_src")
            lt_rp  = get_lt_jours(src_rp)
            st.info(f"Délai auto : **{lt_rp}j**")
            rd  = st.number_input("Demande moy. (u/j)", value=50.0, step=1.0,  key="rp_rd")
        with c2:
            rss = st.number_input("Stock sécu. (u)",    value=132.0, step=10.0, key="rp_rss")
            rcu = st.number_input("Stock actuel (u)",   value=500.0, step=10.0, key="rp_rcu")
        with c3:
            rdm = st.number_input("Demande max (u/j)",  value=70.0,  step=1.0,  key="rp_rdm")

        PR = rd * lt_rp + rss
        dL = (rcu - rss) / rd if rd > 0 else 0
        if rcu <= rss:
            st.markdown("<div class='alert-critical'><span class='alert-icon'><i class='bi bi-exclamation-octagon-fill'></i></span><div>RUPTURE IMMINENTE — Commander immédiatement !</div></div>", unsafe_allow_html=True)
        elif rcu <= PR:
            st.markdown("<div class='alert-warning'><span class='alert-icon'><i class='bi bi-exclamation-triangle-fill'></i></span><div>COMMANDER MAINTENANT — Point de réappro. atteint</div></div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div class='alert-ok'><span class='alert-icon'><i class='bi bi-check-circle-fill'></i></span><div>Stock suffisant — {fmtInt(rcu)} u &gt; PR ({fmtInt(PR)} u)</div></div>", unsafe_allow_html=True)

        c = st.columns(3)
        for i, (l, v, u, col) in enumerate([
            ("Point de réappro.", fmtInt(PR),          "unités",              C_GREEN),
            ("Jours restants",    fmt(dL, 1),           "jours",               C_MUTED),
            ("Délai critique",    fmt(dL - lt_rp, 1),  "jours avant urgence", C_GOLD),
        ]):
            with c[i]: st.markdown(mcard(l, v, u, color=col), unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE : ALERTES — chargement auto depuis calcul, sinon import XLSX
# ═══════════════════════════════════════════════════════════════════════════════
def alertes():
    st.markdown('<h2><i class="bi bi-bell-fill" style="color:#dc3535;margin-right:10px"></i>Alertes &amp; Surveillance des stocks</h2>', unsafe_allow_html=True)

    def clean_num(v):
        if v is None: return 0.0
        s = str(v).strip().replace(" ", "").replace(",", ".")
        if s in ("#DIV/0!", "#N/A", "#VALEUR!", "#REF!", "", "-", "—"): return 0.0
        try:   return float(s)
        except: return 0.0

    df_alerte_all = {}

    # ── Détection automatique des données de session ──────────────────────────
    has_session = ("df_resultats" in st.session_state and
                   bool(st.session_state["df_resultats"]))

    if has_session:
        # Afficher un bandeau de confirmation bien visible
        n_arts = sum(len(v) for v in st.session_state["df_resultats"].values())
        feuilles = list(st.session_state["df_resultats"].keys())


        src_choice = st.radio(
            "Source des données pour les alertes",
            ["Utiliser les indicateurs du calcul automatique", "Importer un fichier XLSX"],
            horizontal=True, key="alerte_src_radio")
    else:
        src_choice = "Importer un fichier XLSX"
        st.info("Aucun calcul automatique en cours de session. Importez un fichier XLSX ou lancez d'abord un Calcul automatique.")

    # ── Chargement depuis session ─────────────────────────────────────────────
    if has_session and "calcul automatique" in src_choice.lower():
        for sheet_name, df_calc in st.session_state["df_resultats"].items():
            df_tmp = df_calc.copy()
            # S'assurer que la colonne SS a le bon nom
            if "Stock sécurité" in df_tmp.columns and "Stock sécurité calculé" not in df_tmp.columns:
                df_tmp["Stock sécurité calculé"] = df_tmp["Stock sécurité"]
            df_alerte_all[sheet_name] = df_tmp

    # ── Chargement depuis fichier XLSX ────────────────────────────────────────
    else:
        uploaded = st.file_uploader(
            "Importez le fichier XLSX exporté depuis le calcul automatique",
            type=["xlsx", "xls"], key="alertes_upload")

        if uploaded:
            try:
                all_raw     = pd.read_excel(uploaded, sheet_name=None)
                sheet_names = list(all_raw.keys())

                if len(sheet_names) > 1:
                    st.info(f"**{len(sheet_names)} feuilles détectées :** {', '.join(sheet_names)}")
                    feuille_alerte = st.radio("Feuille à analyser", sheet_names, horizontal=True, key="alerte_sheet_sel")
                    sheets_proc    = [feuille_alerte]
                else:
                    sheets_proc = sheet_names

                for sn in sheets_proc:
                    df_raw = all_raw[sn]
                    df_raw.columns = [str(c).strip() for c in df_raw.columns]
                    df_raw = df_raw.dropna(how="all")

                    cmap = {}
                    for col in df_raw.columns:
                        cl = col.lower().strip()
                        if "code" in cl:                                           cmap["Code article"] = col
                        elif "description" in cl or "désignation" in cl:          cmap["Description Article"] = col
                        elif "source" in cl:                                       cmap["Source"] = col
                        elif "stock sécurité" in cl or "ss (u)" in cl:            cmap["Stock sécurité calculé"] = col
                        elif "stock actuel" in cl:                                 cmap["Stock actuel"] = col
                        elif "point de commande" in cl or "point commande" in cl: cmap["Point de commande"] = col
                        elif "commande en cours" in cl or "cmd en cours" in cl:   cmap["Commande en cours"] = col
                        elif "consommation" in cl or "conso" in cl:               cmap["Conso mois moy"] = col

                    df_m = df_raw.rename(columns={v: k for k, v in cmap.items()})
                    for c in ["Stock sécurité calculé","Stock actuel","Point de commande",
                               "Commande en cours","Conso mois moy"]:
                        if c in df_m.columns:
                            df_m[c] = df_m[c].apply(clean_num)

                    if "Code article" in df_m.columns:
                        df_m = df_m[df_m["Code article"].notna()]
                    df_alerte_all[sn] = df_m
                    st.success(f"Feuille **{sn}** : {len(df_m)} articles chargés")
            except Exception as e:
                st.error(f"Erreur lecture fichier : {e}")

    if not df_alerte_all:
        return

    # ═══════════════════════════════════════════════════════════════════════════
    # TRAITEMENT DES ALERTES PAR FEUILLE
    # ═══════════════════════════════════════════════════════════════════════════
    for sheet_name, df in df_alerte_all.items():
        if len(df_alerte_all) > 1:
            st.markdown(f'<hr/><h3><i class="bi bi-file-earmark-spreadsheet" style="color:#2d6a2d"></i> {sheet_name}</h3>', unsafe_allow_html=True)

        df = df.copy()
        ss_col  = "Stock sécurité calculé" if "Stock sécurité calculé" in df.columns else "Stock sécurité"
        stk_col = "Stock actuel"
        pr_col  = "Point de commande"

        for col_needed in [ss_col, stk_col, pr_col]:
            if col_needed not in df.columns:
                df[col_needed] = 0

        # ── Mise à jour du stock actuel si les données viennent de la session ─
        st.markdown('<h4><i class="bi bi-pencil-square" style="color:#2d6a2d;margin-right:8px"></i>Mettre à jour le stock actuel</h4>', unsafe_allow_html=True)
        st.caption("Les stocks actuel et commandes en cours sont issus du fichier importé — modifiez-les si nécessaire.")

        # Tableau interactif de mise à jour des stocks
        stock_update_data = []
        for _, row in df.iterrows():
            stock_update_data.append({
                "Code article":       str(row.get("Code article", "")),
                "Description":        str(row.get("Description Article", ""))[:35],
                "Stock actuel (u)":   float(clean_num(row.get(stk_col, 0))),
                "Commande en cours (u)": float(clean_num(row.get("Commande en cours", 0))),
            })

        df_stock_edit = pd.DataFrame(stock_update_data)
        edited_stocks = st.data_editor(
            df_stock_edit,
            use_container_width=True,
            hide_index=True,
            key=f"stock_edit_{sheet_name}",
            column_config={
                "Code article":          st.column_config.TextColumn("Code", disabled=True, width="small"),
                "Description":           st.column_config.TextColumn("Description", disabled=True, width="large"),
                "Stock actuel (u)":      st.column_config.NumberColumn("Stock actuel (u)", min_value=0, step=1, format="%d"),
                "Commande en cours (u)": st.column_config.NumberColumn("Commande en cours (u)", min_value=0, step=1, format="%d"),
            }
        )

        # Réinjecter les stocks mis à jour
        code_to_stk = dict(zip(edited_stocks["Code article"].astype(str), edited_stocks["Stock actuel (u)"]))
        code_to_cmd = dict(zip(edited_stocks["Code article"].astype(str), edited_stocks["Commande en cours (u)"]))
        df[stk_col]             = df["Code article"].astype(str).map(code_to_stk).fillna(0)
        df["Commande en cours"] = df["Code article"].astype(str).map(code_to_cmd).fillna(0)

        # ── Calcul des alertes ────────────────────────────────────────────────
        def get_alerte(row):
            stk = clean_num(row.get(stk_col, 0))
            ss  = clean_num(row.get(ss_col, 0))
            pr  = clean_num(row.get(pr_col, 0))
            cmd = clean_num(row.get("Commande en cours", 0))
            if stk <= 0 or stk < ss * 0.5:
                return "Rupture critique"
            elif stk < ss:
                return "SS dépassé"
            elif stk <= pr and not (stk + cmd > pr):
                return "Commander maintenant"
            else:
                return "Normal"

        df["Alerte"] = df.apply(get_alerte, axis=1)

        rupt    = df[df["Alerte"] == "Rupture critique"]
        ss_dep  = df[df["Alerte"] == "SS dépassé"]
        cmd_now = df[df["Alerte"] == "Commander maintenant"]
        ok      = df[df["Alerte"] == "Normal"]

        st.divider()
        c1,c2,c3,c4 = st.columns(4)
        with c1: st.markdown(mcard("Rupture critique",     str(len(rupt)),    "articles", color=C_RED,     icon="bi-x-octagon-fill"),           unsafe_allow_html=True)
        with c2: st.markdown(mcard("SS dépassé",           str(len(ss_dep)),  "articles", color="#e65100", icon="bi-exclamation-triangle-fill"), unsafe_allow_html=True)
        with c3: st.markdown(mcard("Commander maintenant", str(len(cmd_now)), "articles", color=C_GOLD,    icon="bi-cart-plus-fill"),            unsafe_allow_html=True)
        with c4: st.markdown(mcard("Normal",               str(len(ok)),      "articles", color=C_GREEN2,  icon="bi-check-circle"),              unsafe_allow_html=True)

        st.divider()

        # ── Tableau filtrable ─────────────────────────────────────────────────
        filtre = st.multiselect(
            "Filtrer par niveau d'alerte",
            options=["Rupture critique","SS dépassé","Commander maintenant","Normal"],
            default=["Rupture critique","SS dépassé","Commander maintenant"],
            key=f"filtre_{sheet_name}")

        df_show = df[df["Alerte"].isin(filtre)] if filtre else df

        disp_cols = ["Code article","Description Article","Source"]
        for c in [ss_col, stk_col, pr_col, "Commande en cours", "Alerte"]:
            if c in df_show.columns:
                disp_cols.append(c)

        st.dataframe(df_show[[c for c in disp_cols if c in df_show.columns]],
                     use_container_width=True, hide_index=True)



        # ── Actions prioritaires ──────────────────────────────────────────────
        st.divider()
        st.markdown('<h4><i class="bi bi-lightning-charge-fill" style="color:#dc3535;margin-right:8px"></i>Actions prioritaires</h4>', unsafe_allow_html=True)

        if len(rupt) == 0 and len(ss_dep) == 0 and len(cmd_now) == 0:
            st.markdown("<div class='alert-ok'><span class='alert-icon'><i class='bi bi-check-circle-fill'></i></span><div>Aucune alerte — tous les stocks sont suffisants.</div></div>", unsafe_allow_html=True)

        for _, r in rupt.iterrows():
            stk = clean_num(r.get(stk_col, 0)); ss = clean_num(r.get(ss_col, 0))
            st.markdown(
                f"<div class='alert-critical'><span class='alert-icon'><i class='bi bi-x-octagon-fill'></i></span>"
                f"<div><strong>{r.get('Code article','')} — {r.get('Description Article','')}</strong><br/>"
                f"<span style='font-size:.85rem'>Stock actuel : <strong>{stk:,.0f}</strong> u | SS : {ss:,.0f} u | "
                f"Écart : <strong>{stk-ss:,.0f}</strong> u — Commander immédiatement !</span></div></div>",
                unsafe_allow_html=True)

        for _, r in ss_dep.iterrows():
            stk = clean_num(r.get(stk_col, 0)); ss = clean_num(r.get(ss_col, 0))
            pr  = clean_num(r.get(pr_col, 0))
            st.markdown(
                f"<div class='alert-warning'><span class='alert-icon'><i class='bi bi-exclamation-triangle-fill'></i></span>"
                f"<div><strong>{r.get('Code article','')} — {r.get('Description Article','')}</strong><br/>"
                f"<span style='font-size:.85rem'>Stock : {stk:,.0f} u | SS : {ss:,.0f} u | PR : {pr:,.0f} u — Stock de sécurité entamé</span></div></div>",
                unsafe_allow_html=True)

        for _, r in cmd_now.iterrows():
            stk = clean_num(r.get(stk_col, 0)); pr = clean_num(r.get(pr_col, 0))
            lt  = get_lt_jours(r.get("Source", "")) if r.get("Source") else "?"
            st.markdown(
                f"<div class='alert-info'><span class='alert-icon'><i class='bi bi-info-circle-fill'></i></span>"
                f"<div><strong>{r.get('Code article','')} — {r.get('Description Article','')}</strong><br/>"
                f"<span style='font-size:.85rem'>Stock : {stk:,.0f} u ≤ PR : {pr:,.0f} u | Délai : {lt}j — Lancer la commande maintenant</span></div></div>",
                unsafe_allow_html=True)





# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    URL_LOGO = "https://raw.githubusercontent.com/Ghofrane13/medoil-supply-app/main/logo.png"
    st.markdown(f"""
    <div style="display:flex;align-items:center;padding:.3rem 0 .5rem">
        <img src={URL_LOGO} style="width:44px;margin-right:10px;border-radius:6px">
        <div>
            <div style="color:#ffffff;font-weight:700;font-size:1.15rem;line-height:1.1">Med oil</div>
            <div style="color:#f5c518;font-size:.65rem;letter-spacing:.1em;text-transform:uppercase">Supply Chain · v3.0</div>
        </div>
    </div>""", unsafe_allow_html=True)
    st.divider()

    pg = st.navigation([
        st.Page(accueil,         title="Accueil",            icon=":material/home:"),
        st.Page(import_calcul,   title="Calcul automatique", icon=":material/upload_file:"),
        st.Page(calculateurs,    title="Calculateurs SC",     icon=":material/calculate:"),
        st.Page(alertes,         title="Alertes & Stocks",    icon=":material/notifications:"),
      
    ])

    st.markdown("<p style='color:#4a6a4a;font-size:.72rem;'>© 2025 Medoil — Tous droits réservés</p>",
                unsafe_allow_html=True)

pg.run()
